"""
AI Knowledge Centre - Document Parsers
Uses Unstructured.io for document parsing with OCR fallback for scanned PDFs.
"""

import os
import logging
from typing import List, Dict, Any

logger = logging.getLogger(__name__)


def parse_document(file_path: str) -> List[Dict[str, Any]]:
    """
    Parse a document and extract content with metadata.

    Args:
        file_path: Path to the document file

    Returns:
        List of sections with text, page, section metadata
    """
    file_ext = os.path.splitext(file_path)[1].lower()

    try:
        if file_ext == ".pdf":
            return _parse_pdf(file_path)
        elif file_ext in (".docx", ".doc"):
            return _parse_docx(file_path)
        elif file_ext in (".xlsx", ".xls"):
            return _parse_excel(file_path)
        elif file_ext in (".txt", ".md", ".csv"):
            return _parse_text(file_path)
        else:
            return _parse_with_unstructured(file_path)
    except Exception as e:
        logger.error(f"Parsing failed for {file_path}: {e}")
        # Try OCR fallback for PDFs
        if file_ext == ".pdf":
            logger.info("Attempting OCR fallback for scanned PDF")
            return _parse_pdf_ocr(file_path)
        return []


def _parse_pdf(file_path: str) -> List[Dict[str, Any]]:
    """Parse PDF using Unstructured.io."""
    try:
        from unstructured.partition.auto import partition

        elements = partition(filename=file_path)
        sections = []

        for i, element in enumerate(elements):
            text = str(element).strip()
            if text:
                sections.append({
                    "text": text,
                    "page": getattr(element.metadata, "page_number", 0) or 0,
                    "section": getattr(element.metadata, "section_header", "") or "",
                    "filename": os.path.basename(file_path),
                    "element_type": type(element).__name__,
                })

        return sections
    except Exception as e:
        logger.warning(f"Unstructured PDF parsing failed: {e}")
        return []


def _parse_pdf_ocr(file_path: str) -> List[Dict[str, Any]]:
    """Parse scanned PDF using OCR (pdf2image + pytesseract)."""
    try:
        from pdf2image import convert_from_path
        import pytesseract

        images = convert_from_path(file_path, dpi=300)
        sections = []

        for page_num, image in enumerate(images, 1):
            text = pytesseract.image_to_string(image)
            text = text.strip()
            if text:
                sections.append({
                    "text": text,
                    "page": page_num,
                    "section": "",
                    "filename": os.path.basename(file_path),
                    "element_type": "OCR",
                })

        return sections
    except Exception as e:
        logger.error(f"OCR parsing failed: {e}")
        return []


def _parse_docx(file_path: str) -> List[Dict[str, Any]]:
    """Parse DOCX files using python-docx."""
    try:
        from docx import Document

        doc = Document(file_path)
        sections = []
        current_section = ""
        current_text = []

        for para in doc.paragraphs:
            # Detect section headers (Heading styles)
            if para.style.name.startswith("Heading"):
                # Save previous section
                if current_text:
                    sections.append({
                        "text": "\n".join(current_text),
                        "page": 0,
                        "section": current_section,
                        "filename": os.path.basename(file_path),
                        "element_type": "paragraph",
                    })
                current_section = para.text.strip()
                current_text = []
            elif para.text.strip():
                current_text.append(para.text.strip())

        # Save last section
        if current_text:
            sections.append({
                "text": "\n".join(current_text),
                "page": 0,
                "section": current_section,
                "filename": os.path.basename(file_path),
                "element_type": "paragraph",
            })

        return sections
    except Exception as e:
        logger.warning(f"DOCX parsing failed: {e}")
        return []


def _parse_excel(file_path: str) -> List[Dict[str, Any]]:
    """Parse Excel files, treating each row as a separate section."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        sections = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                continue

            # Use first row as header
            headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]

            for row_idx, row in enumerate(rows[1:], 1):
                row_text_parts = []
                for col_idx, cell in enumerate(row):
                    if cell is not None:
                        header = headers[col_idx] if col_idx < len(headers) else f"col_{col_idx}"
                        row_text_parts.append(f"{header}: {cell}")

                if row_text_parts:
                    sections.append({
                        "text": " | ".join(row_text_parts),
                        "page": 0,
                        "section": f"Sheet: {sheet_name}, Row: {row_idx}",
                        "filename": os.path.basename(file_path),
                        "element_type": "excel_row",
                    })

        wb.close()
        return sections
    except Exception as e:
        logger.warning(f"Excel parsing failed: {e}")
        return []


def _parse_text(file_path: str) -> List[Dict[str, Any]]:
    """Parse plain text files."""
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()

        if not content.strip():
            return []

        return [{
            "text": content.strip(),
            "page": 0,
            "section": os.path.basename(file_path),
            "filename": os.path.basename(file_path),
            "element_type": "text",
        }]
    except Exception as e:
        logger.warning(f"Text parsing failed: {e}")
        return []


def _parse_with_unstructured(file_path: str) -> List[Dict[str, Any]]:
    """Generic parser using Unstructured.io for unknown file types."""
    try:
        from unstructured.partition.auto import partition

        elements = partition(filename=file_path)
        sections = []

        for element in elements:
            text = str(element).strip()
            if text:
                sections.append({
                    "text": text,
                    "page": getattr(element.metadata, "page_number", 0) or 0,
                    "section": getattr(element.metadata, "section_header", "") or "",
                    "filename": os.path.basename(file_path),
                    "element_type": type(element).__name__,
                })

        return sections
    except Exception as e:
        logger.warning(f"Unstructured parsing failed for {file_path}: {e}")
        return []
